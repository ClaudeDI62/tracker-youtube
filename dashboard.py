"""YouTube Recommendation Tracker — Dashboard Streamlit."""

import io
import re
from datetime import date, datetime

import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from src.config import SUPABASE_URL, SUPABASE_KEY
from supabase import create_client

# --- Auth ---
def check_password():
    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False

    if st.session_state.authenticated:
        return True

    st.title("YouTube Recommendation Tracker")
    st.markdown("Inserisci la password per accedere.")
    pwd = st.text_input("Password", type="password")
    if st.button("Accedi"):
        if pwd == st.secrets.get("dashboard_password", "tracker2026"):
            st.session_state.authenticated = True
            st.rerun()
        else:
            st.error("Password errata.")
    return False


# --- DB ---
@st.cache_resource
def get_db():
    return create_client(SUPABASE_URL, SUPABASE_KEY)


def query(table, select="*", filters=None, order=None):
    q = get_db().table(table).select(select)
    if filters:
        for method, args in filters:
            q = getattr(q, method)(*args)
    if order:
        q = q.order(order[0], desc=order[1])
    return q.execute().data


# --- Helpers ---
def fmt_pct(val):
    if val is None:
        return "—"
    v = float(val)
    sign = "+" if v >= 0 else ""
    return f"{sign}{v:.1f}%"


def fmt_price(val):
    if val is None:
        return "—"
    return f"${float(val):,.2f}"


def days_since(date_str):
    if not date_str:
        return "—"
    d = date.fromisoformat(str(date_str)[:10])
    return (date.today() - d).days


# --- Page: Classifica Canali ---
def page_rankings():
    st.header("Classifica Canali")
    st.caption("Il rating misura l'accuratezza storica, non predice quella futura.")

    ratings = query("ratings", "*, channels(name)", order=("score", True))

    if not ratings:
        st.info("Nessun rating disponibile ancora. Servono almeno 7 giorni di dati.")
        return

    horizon = st.radio("Orizzonte", [7, 30, 90], format_func=lambda x: f"{x} giorni", horizontal=True)

    filtered = [r for r in ratings if r["horizon_days"] == horizon]
    filtered.sort(key=lambda r: (float(r["score"] or 0), float(r["avg_excess_return"] or 0)), reverse=True)

    if not filtered:
        st.info(f"Nessun dato per l'orizzonte a {horizon} giorni.")
        return

    rows = []
    for i, r in enumerate(filtered):
        ch_name = r.get("channels", {}).get("name", "?")
        medal = ["🥇", "🥈", "🥉"][i] if i < 3 else f"#{i+1}"
        rows.append({
            "Pos.": medal,
            "Canale": ch_name,
            "N. Reco": r["n_recos"],
            "Hit Rate": f"{float(r['hit_rate'] or 0) * 100:.0f}%",
            "Wilson": f"{float(r['score'] or 0):.3f}",
            "Excess Medio": fmt_pct(r["avg_excess_return"]),
        })

    st.dataframe(rows, use_container_width=True, hide_index=True)

    st.divider()
    for i, r in enumerate(filtered):
        ch_name = r.get("channels", {}).get("name", "?")
        n = r["n_recos"]
        hit_rate = float(r["hit_rate"] or 0) * 100
        wilson = float(r["score"] or 0)
        excess = float(r["avg_excess_return"] or 0)

        medal = ["🥇", "🥈", "🥉"][i] if i < 3 else f"#{i+1}"
        st.subheader(f"{medal} {ch_name}")
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("N. Reco", n)
        c2.metric("Hit Rate", f"{hit_rate:.0f}%")
        c3.metric("Wilson", f"{wilson:.3f}")
        c4.metric("Excess", fmt_pct(excess))


# --- Page: Raccomandazioni Aperte ---
def page_open_recos():
    st.header("Raccomandazioni Aperte")

    recos = query(
        "recommendations",
        "*, videos!inner(title, channels!inner(name))",
        filters=[("in_", ("status", ["open", "evaluated"]))],
        order=("reco_date", True),
    )

    if not recos:
        st.info("Nessuna raccomandazione aperta.")
        return

    st.caption(f"{len(recos)} raccomandazioni aperte")

    action_filter = st.multiselect("Filtra per azione", ["buy", "sell", "hold"], default=["buy", "sell", "hold"])

    for r in recos:
        if r["action"] not in action_filter:
            continue

        channel_name = r.get("videos", {}).get("channels", {}).get("name", "?")
        video_title = r.get("videos", {}).get("title", "?")
        days = days_since(r["reco_date"])

        action_colors = {"buy": "🟢", "sell": "🔴", "hold": "🟡"}
        action_icon = action_colors.get(r["action"], "⚪")

        with st.container(border=True):
            col1, col2, col3, col4 = st.columns([3, 1, 1, 1])
            with col1:
                st.markdown(f"**{action_icon} {r['action'].upper()} — {r['asset_name']}**")
                st.caption(f"{channel_name} · {r['reco_date']} · {days} giorni fa")
            with col2:
                st.markdown(f"**Ticker:** {r['ticker'] or '—'}")
            with col3:
                st.markdown(f"**Prezzo entrata:** {fmt_price(r['price_at_reco'])}")
            with col4:
                if r["conditional"]:
                    st.markdown("⚠️ Condizionale")
                if r["target_price"]:
                    st.markdown(f"**Target:** {fmt_price(r['target_price'])}")

            if r.get("rationale"):
                st.caption(f"Motivazione: {r['rationale']}")
            if r.get("quote"):
                st.caption(f'Citazione: "{r["quote"]}"')


# --- Page: Dettaglio Canale ---
def page_channel_detail():
    st.header("Dettaglio Canale")

    channels = query("channels", "id, name", filters=[("eq", ("active", True))])
    if not channels:
        st.info("Nessun canale attivo.")
        return

    ch_names = {c["name"]: c["id"] for c in channels}
    selected = st.selectbox("Seleziona canale", list(ch_names.keys()))
    ch_id = ch_names[selected]

    # Channel ratings
    ratings = query("ratings", "*", filters=[("eq", ("channel_id", ch_id))])
    if ratings:
        st.subheader("Rating")
        cols = st.columns(3)
        for i, h in enumerate([7, 30, 90]):
            r = next((x for x in ratings if x["horizon_days"] == h), None)
            with cols[i]:
                if r:
                    hit_rate = float(r["hit_rate"] or 0) * 100
                    st.metric(
                        f"{h} giorni",
                        f"{hit_rate:.0f}% ({r['n_recos']} reco)",
                        f"Wilson: {float(r['score'] or 0):.3f}",
                    )
                else:
                    st.metric(f"{h} giorni", "—", "Dati insufficienti")

    # Channel recommendations
    st.subheader("Storico Raccomandazioni")

    videos = query("videos", "id, title", filters=[("eq", ("channel_id", ch_id))])
    video_ids = [v["id"] for v in videos]

    if not video_ids:
        st.info("Nessun video elaborato per questo canale.")
        return

    recos = query(
        "recommendations", "*",
        filters=[("in_", ("video_id", video_ids))],
        order=("reco_date", True),
    )

    if not recos:
        st.info("Nessuna raccomandazione estratta per questo canale.")
        return

    reco_ids = [r["id"] for r in recos]
    evals = query(
        "evaluations", "*",
        filters=[("in_", ("recommendation_id", reco_ids))],
    )
    evals_by_reco = {}
    for e in evals:
        evals_by_reco.setdefault(e["recommendation_id"], []).append(e)

    action_colors = {"buy": "🟢", "sell": "🔴", "hold": "🟡"}

    for r in recos:
        icon = action_colors.get(r["action"], "⚪")
        with st.container(border=True):
            st.markdown(f"**{icon} {r['action'].upper()} {r['asset_name']}** ({r['ticker'] or '—'}) — {r['reco_date']}")

            if r.get("rationale"):
                st.caption(r["rationale"])

            my_evals = evals_by_reco.get(r["id"], [])
            if my_evals:
                eval_cols = st.columns(len(my_evals))
                for j, ev in enumerate(sorted(my_evals, key=lambda x: x["horizon_days"])):
                    with eval_cols[j]:
                        hit = ev["hit"]
                        if hit is None:
                            result = "HOLD"
                            delta_color = "off"
                        elif hit:
                            result = "HIT ✓"
                            delta_color = "normal"
                        else:
                            result = "MISS ✗"
                            delta_color = "inverse"

                        st.metric(
                            f"{ev['horizon_days']}g",
                            result,
                            fmt_pct(ev["excess_return_pct"]),
                            delta_color=delta_color,
                        )


# --- Page: Download Excel ---
def page_excel():
    st.header("Download Excel")
    st.markdown("Scarica tutti i dati in un file Excel con tre fogli: Recommendations, Evaluations, Ratings.")

    if st.button("Genera Excel", type="primary"):
        with st.spinner("Generazione in corso..."):
            excel_data = generate_excel()
        st.download_button(
            "Scarica il file",
            data=excel_data,
            file_name=f"tracker_youtube_{date.today().isoformat()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


def generate_excel():
    wb = openpyxl.Workbook()

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        bottom=Side(style="thin", color="D8DDE4"),
    )

    def style_header(ws, cols):
        for i, col in enumerate(cols, 1):
            cell = ws.cell(row=1, column=i, value=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

    # Sheet 1: Recommendations
    ws1 = wb.active
    ws1.title = "Recommendations"
    cols1 = ["Channel", "Date", "Action", "Asset", "Ticker", "ISIN", "Type",
             "Price at Reco", "Target Price", "Horizon", "Conditional",
             "Rationale", "Quote", "Status"]
    style_header(ws1, cols1)

    recos = query(
        "recommendations",
        "*, videos!inner(title, channels!inner(name))",
        order=("reco_date", True),
    )
    for r in recos:
        ch = r.get("videos", {}).get("channels", {}).get("name", "")
        ws1.append([
            ch, r["reco_date"], r["action"].upper(), r["asset_name"],
            r["ticker"], r.get("isin"), r["asset_type"],
            float(r["price_at_reco"]) if r["price_at_reco"] else None,
            float(r["target_price"]) if r["target_price"] else None,
            r.get("horizon_text"), r["conditional"],
            r.get("rationale"), r.get("quote"), r["status"],
        ])

    for col in ws1.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws1.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    # Sheet 2: Evaluations
    ws2 = wb.create_sheet("Evaluations")
    cols2 = ["Channel", "Asset", "Ticker", "Action", "Reco Date",
             "Horizon Days", "Asset Return %", "Benchmark Return %",
             "Excess Return %", "Hit"]
    style_header(ws2, cols2)

    evals = query("evaluations", "*, recommendations!inner(*, videos!inner(channels!inner(name)))")
    for e in evals:
        rec = e.get("recommendations", {})
        ch = rec.get("videos", {}).get("channels", {}).get("name", "")
        hit_str = "HIT" if e["hit"] else ("MISS" if e["hit"] is not None else "HOLD")
        ws2.append([
            ch, rec.get("asset_name"), rec.get("ticker"), rec.get("action", "").upper(),
            rec.get("reco_date"), e["horizon_days"],
            float(e["asset_return_pct"]) if e["asset_return_pct"] else None,
            float(e["benchmark_return_pct"]) if e["benchmark_return_pct"] else None,
            float(e["excess_return_pct"]) if e["excess_return_pct"] else None,
            hit_str,
        ])

    for col in ws2.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws2.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    # Sheet 3: Ratings
    ws3 = wb.create_sheet("Ratings")
    cols3 = ["Channel", "Horizon Days", "N Recos", "Hit Rate",
             "Wilson Lower Bound", "Avg Excess Return %", "Score"]
    style_header(ws3, cols3)

    ratings = query("ratings", "*, channels(name)", order=("score", True))
    for r in ratings:
        ch = r.get("channels", {}).get("name", "")
        ws3.append([
            ch, r["horizon_days"], r["n_recos"],
            float(r["hit_rate"]) if r["hit_rate"] else None,
            float(r["wilson_lb"]) if r["wilson_lb"] else None,
            float(r["avg_excess_return"]) if r["avg_excess_return"] else None,
            float(r["score"]) if r["score"] else None,
        ])

    for col in ws3.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws3.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# --- Page: Gestione Canali ---
def _extract_channel_id(url_or_id):
    """Try to extract channel_id from a URL or return as-is if it looks like one."""
    if url_or_id.startswith("UC") and len(url_or_id) == 24:
        return url_or_id
    match = re.search(r"channel/(UC[\w-]{22})", url_or_id)
    if match:
        return match.group(1)
    return url_or_id.strip()


def page_manage_channels():
    st.header("Gestione Canali")
    st.caption("Aggiungi, attiva o disattiva i canali YouTube monitorati.")

    db = get_db()

    channels = db.table("channels").select("*").order("name").execute().data

    # --- Current channels ---
    st.subheader("Canali attuali")
    if channels:
        for ch in channels:
            col1, col2, col3 = st.columns([4, 1, 1])
            with col1:
                status_icon = "🟢" if ch["active"] else "🔴"
                st.markdown(f"**{status_icon} {ch['name']}**")
                st.caption(f"ID: {ch['channel_id']} · Lingua: {ch.get('language', '?')} · Mercato: {ch.get('market_default', '?')}")
            with col2:
                n_videos = len(db.table("videos").select("id").eq("channel_id", ch["id"]).execute().data)
                st.metric("Video", n_videos)
            with col3:
                if ch["active"]:
                    if st.button("Disattiva", key=f"deact_{ch['id']}"):
                        db.table("channels").update({"active": False}).eq("id", ch["id"]).execute()
                        st.rerun()
                else:
                    if st.button("Attiva", key=f"act_{ch['id']}"):
                        db.table("channels").update({"active": True}).eq("id", ch["id"]).execute()
                        st.rerun()
            st.divider()
    else:
        st.info("Nessun canale configurato.")

    # --- Add new channel ---
    st.subheader("Aggiungi nuovo canale")

    with st.form("add_channel", clear_on_submit=True):
        col_a, col_b = st.columns(2)
        with col_a:
            new_name = st.text_input("Nome del canale *", placeholder="Es. Mario Rossi Finance")
            new_url = st.text_input("URL del canale *", placeholder="https://www.youtube.com/@NomeCanale")
        with col_b:
            new_channel_id = st.text_input("Channel ID *", placeholder="UCxxxxxxxxxxxxxxxxxx (inizia con UC, 24 caratteri)")
            lang_col, market_col = st.columns(2)
            with lang_col:
                new_language = st.selectbox("Lingua", ["EN", "IT", "DE", "FR", "ES"])
            with market_col:
                new_market = st.selectbox("Mercato", ["USA", "Europe", "Italy", "Germany", "UK"])

        st.caption("* Campi obbligatori. Il Channel ID si trova nella URL della pagina del canale o cercando 'YouTube Channel ID Finder' su Google.")

        submitted = st.form_submit_button("Aggiungi canale", type="primary")

        if submitted:
            if not new_name or not new_url or not new_channel_id:
                st.error("Compila tutti i campi obbligatori.")
            else:
                clean_id = _extract_channel_id(new_channel_id)
                if not clean_id.startswith("UC") or len(clean_id) != 24:
                    st.error("Il Channel ID deve iniziare con 'UC' e avere 24 caratteri.")
                else:
                    existing = db.table("channels").select("id").eq("channel_id", clean_id).execute()
                    if existing.data:
                        st.error("Questo canale esiste gia nel database.")
                    else:
                        db.table("channels").insert({
                            "name": new_name.strip(),
                            "url": new_url.strip(),
                            "channel_id": clean_id,
                            "language": new_language,
                            "market_default": new_market,
                            "active": True,
                        }).execute()
                        st.success(f"Canale '{new_name}' aggiunto! Sara monitorato dal prossimo job serale.")
                        st.rerun()


# --- Main ---
def main():
    st.set_page_config(
        page_title="YouTube Recommendation Tracker",
        page_icon="📊",
        layout="wide",
    )

    if not check_password():
        return

    st.sidebar.title("📊 YT Tracker")
    st.sidebar.caption(f"Aggiornato al {date.today().strftime('%d/%m/%Y')}")

    page = st.sidebar.radio(
        "Navigazione",
        ["Classifica Canali", "Raccomandazioni Aperte", "Dettaglio Canale", "Download Excel", "Gestione Canali"],
    )

    if page == "Classifica Canali":
        page_rankings()
    elif page == "Raccomandazioni Aperte":
        page_open_recos()
    elif page == "Dettaglio Canale":
        page_channel_detail()
    elif page == "Download Excel":
        page_excel()
    elif page == "Gestione Canali":
        page_manage_channels()


if __name__ == "__main__":
    main()
